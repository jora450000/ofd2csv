#!/usr/bin/python3


##@@# -*- coding: utf-8 -*-
 
from bs4 import BeautifulSoup
import requests as req
import re
import openpyxl

##import _locale, os, sys
##_locale._getdefaultlocale = (lambda *args: ['ru_RU', 'utf8'])


wb=openpyxl.load_workbook(filename='./checks.xlsx')
sheet = wb['Лист1']    
x=2
print("date\tkassa\tsumma\tusluga\tcount\tcost\tsumma_uslugi")  #header of csv

while (True):
  date_check = sheet.cell(row=x, column = 1).value

  if (date_check != None):	
    summa = sheet.cell(row=x, column = 24).value
    kassa = sheet.cell(row=x, column = 3).value
    
    uri_formula = sheet.cell(row=x, column = 26)
    match =re.search(r'"(.*)",', uri_formula.value)
    uri = match.group(1)
    resp = req.get(uri)
    soup = BeautifulSoup(resp.text, 'lxml').body.text.split('chek.pofd.ru')[1].split('N ФН')[0] # cut body of check
#    print("date\tkassa\tsumma\tusluga\tcount\tcost\tsumma_uslugi")  #header of csv
    for p in re.split("(РАСЧЕТА|УСЛУГА|ТОВАР|ПЛАТЕЖ|ИНОЙ ПРЕДМЕТ|Итого без НДС)",soup):
           m = re.split(r'( Сбор на лося до 1 года| Путёвка на лося до 1 года|\d+.\d+|\d+| х )', p)

#           print(f"m==>{m}")
           for j in range(len(m)):
             if (m[j]== '' or m[j] == ' х ' or m[j] == ' '):
                 continue
             elif (m[j][0] == ' ' ):
#                print(f"m={m}")
                try:
                  cost = float(m[j+5+int(m[j+1]=='')])
                  summa = cost * int(m[j+1+int(m[j+1]=='')]) 
                except:
                  continue
                print(f"{date_check}\t {kassa}\t {int(summa)}\t {m[j][1:]}\t {m[j+1]}\t {cost:8.0f}\t {summa:8.0f}")
                j += 5

    x+=1
  else:
    break
